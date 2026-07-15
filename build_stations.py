# -*- coding: utf-8 -*-
"""Sinh libs/stations_hno.js — tọa độ trạm khu vực Hà Nội cho Field Map.

Nguồn (export CCG mới nhất trong Downloads, đổi đường dẫn khi có bản mới):
  - EVCS: 120726_EVCS_Station.xlsx  (sheet "Danh sách trạm sạc 1")
  - BSS : 120726_BSS_Station.xlsx   (sheet "Danh sách trạm đổi pin")

Lọc: Tỉnh Thành chứa "Hà Nội" và có đủ lat/long hợp lệ.
Output: const STATIONS = {"C.HNO01031":[21.03,105.85,"E","Tên trạm","Phường/Xã"],...}
        ("E" = trạm sạc EVCS, "B" = tủ đổi pin BSS)

Chạy lại khi có file trạm mới:  py build_stations.py [file_evcs] [file_bss]
"""
import json
import os
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

SRC_EVCS = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\Lenovo\Downloads\120726_EVCS_Station.xlsx"
SRC_BSS = sys.argv[2] if len(sys.argv) > 2 else r"C:\Users\Lenovo\Downloads\120726_BSS_Station.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "libs", "stations_hno.js")

# Hà Nội: lat 20.5–21.5, lng 105.2–106.1 (chặn tọa độ rác)
LAT_MIN, LAT_MAX, LNG_MIN, LNG_MAX = 20.3, 21.7, 105.0, 106.3


def load(path, kind):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}
    i_code = idx["Mã Trạm"]
    i_name = idx["Tên trạm"]
    i_lat = idx["Vĩ độ (lat)"]
    i_lng = idx["Kinh độ (long)"]
    # Cột "(Mới)" trong export 12/07 bị lệch dữ liệu (chứa tên phường) -> dùng cột cũ
    i_prov = idx["Tỉnh Thành"]
    i_ward = idx.get("Phường/Xã")
    out = {}
    n_hn = n_badcoord = 0
    for r in rows:
        code = str(r[i_code]).strip() if r[i_code] else ""
        if not code or code == "None":
            continue
        prov = str(r[i_prov]).strip() if r[i_prov] else ""
        if "Hà Nội" not in prov:
            continue
        n_hn += 1
        try:
            lat = float(str(r[i_lat]).replace(",", "."))
            lng = float(str(r[i_lng]).replace(",", "."))
        except (TypeError, ValueError):
            n_badcoord += 1
            continue
        if not (LAT_MIN <= lat <= LAT_MAX and LNG_MIN <= lng <= LNG_MAX):
            n_badcoord += 1
            continue
        name = str(r[i_name]).strip()[:60] if r[i_name] else ""
        ward = str(r[i_ward]).strip() if (i_ward is not None and r[i_ward]) else ""
        out[code] = [round(lat, 6), round(lng, 6), kind, name, ward]
    print("%s: %d trạm Hà Nội, %d thiếu/sai tọa độ -> giữ %d" % (kind, n_hn, n_badcoord, len(out)))
    return out


stations = load(SRC_EVCS, "E")
stations.update(load(SRC_BSS, "B"))

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("// Sinh tự động bởi build_stations.py — KHÔNG sửa tay\n")
    f.write("// %s + %s\n" % (os.path.basename(SRC_EVCS), os.path.basename(SRC_BSS)))
    f.write("const STATIONS = ")
    f.write(json.dumps(stations, ensure_ascii=False, separators=(",", ":")))
    f.write(";\n")
print("Đã ghi %s: %d trạm (%.0f KB)" % (OUT, len(stations), os.path.getsize(OUT) / 1024))
